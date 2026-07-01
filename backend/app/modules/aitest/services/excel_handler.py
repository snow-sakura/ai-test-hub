"""
Excel 导入导出处理器

提供测试用例的 Excel 导入导出功能，使用 openpyxl 实现。
12 列格式，含优先级着色，支持 Markdown 格式用例解析。
"""

import io
import logging
import re

logger = logging.getLogger(__name__)

# ======================================================================
# 优先级颜色映射
# ======================================================================

PRIORITY_COLORS = {
    "p0": "FF4D4F",  # 红色
    "p1": "FF7A00",  # 橙色
    "p2": "1890FF",  # 蓝色
    "p3": "D9D9D9",  # 灰色
}

HEADERS = [
    "用例标题", "模块", "优先级", "测试类型",
    "前置条件", "测试步骤", "预期结果", "状态",
    "标签", "版本", "评审人", "备注",
]

EXPECTED_HEADERS = HEADERS


def export_cases_to_xlsx(cases: list, project_name: str = "测试用例") -> bytes:
    """
    导出测试用例列表为 Excel 字节流。

    参数:
        cases: list[TestCase] — SQLAlchemy ORM 对象列表
        project_name: str — 项目名称（用于工作表名）

    返回:
        bytes — .xlsx 文件的字节内容
    """
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = project_name[:31]  # 工作表名限制 31 字符

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    for row_idx, case in enumerate(cases, 2):
        priority = (case.priority or "p2").lower()
        color = PRIORITY_COLORS.get(priority, "D9D9D9")
        row_data = [
            case.name,
            case.module or "",
            priority,
            case.test_type or "functional",
            case.precondition or "",
            case.test_steps or "",
            case.expected_result or "",
            case.status or "draft",
            ", ".join(case.tags) if hasattr(case, "tags") and case.tags else "",
            str(case.version_id or ""),
            "",
            "",
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            # 优先级列着色
            if col == 3:  # 优先级列
                cell.font = Font(color=color if priority != "p2" else "000000", bold=True)
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid",
                ) if priority != "p2" else PatternFill()

    # 列宽
    col_widths = [30, 15, 10, 12, 30, 40, 30, 10, 20, 10, 10, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def parse_xlsx_cases(file_bytes: bytes) -> list[dict]:
    """
    解析 Excel 文件中的测试用例数据。

    返回:
        list[dict] — 每个 dict 包含 TestCase 的字段数据
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if ws is None:
        return []

    # 读取表头映射
    header_map = {}
    for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), 1):
        if cell and str(cell).strip() in EXPECTED_HEADERS:
            header_map[str(cell).strip()] = col_idx - 1  # 0-indexed

    get_col = lambda row, name: str(row[header_map[name]]).strip() if name in header_map and len(row) > header_map[name] else ""
    if "用例标题" not in header_map:
        return []

    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = get_col(row, "用例标题")
        if not title:
            continue
        priority = get_col(row, "优先级").lower()
        if priority not in ("p0", "p1", "p2", "p3"):
            priority = "p2"

        status = get_col(row, "状态").lower()
        if status not in ("draft", "active", "deprecated"):
            status = "draft"

        test_type = get_col(row, "测试类型").lower()
        if not test_type:
            test_type = "functional"

        tags_str = get_col(row, "标签")
        tags = [t.strip() for t in tags_str.split(",") if t.strip()] if tags_str else None

        cases.append({
            "name": title,
            "module": get_col(row, "模块") or None,
            "priority": priority,
            "test_type": test_type,
            "precondition": get_col(row, "前置条件") or None,
            "test_steps": get_col(row, "测试步骤") or None,
            "expected_result": get_col(row, "预期结果") or None,
            "status": status,
            "tags": tags,
        })

    return cases


def parse_markdown_cases(markdown_text: str) -> list[dict]:
    """
    解析 AI 生成的 Markdown 格式测试用例。

    支持格式：
    ```
    ## TC001
    **标题**: 登录功能测试
    **优先级**: P1
    ...

    或 --- 分隔的用例
    ```

    返回:
        list[dict] — TestCase 兼容的字典列表
    """
    cases = []
    # 按 --- 或 ## TC 分割用例块
    blocks = re.split(r'\n---+\n|\n#{2,3}\s+TC\d+', markdown_text)

    for block in blocks:
        block = block.strip()
        if not block:
            continue

        case = _parse_single_markdown_case(block)
        if case and case.get("name"):
            cases.append(case)

    return cases


def _parse_single_markdown_case(text: str) -> dict | None:
    """解析单个 Markdown 用例块"""
    result = {
        "name": "",
        "module": None,
        "priority": "p2",
        "test_type": "functional",
        "precondition": None,
        "test_steps": None,
        "expected_result": None,
        "status": "active",
        "source": "ai_generated",
        "tags": None,
    }

    # 按行提取字段
    lines = text.split("\n")
    steps_lines = []
    expected_lines = []
    in_steps = False
    in_expected = False
    other_sections = []

    for line in lines:
        stripped = line.strip()
        lower = stripped.lower()

        # 尝试匹配 "**字段名**: 值" 格式
        field_match = re.match(r'\*{0,2}(标题|名称|用例名称|用例标题|模块|优先级|测试类型|前置条件|状态)\*{0,2}\s*[:：]\s*(.+)', stripped)
        if field_match:
            key, value = field_match.group(1), field_match.group(2).strip()
            if key in ("标题", "名称", "用例名称", "用例标题"):
                result["name"] = value
            elif key == "模块":
                result["module"] = value
            elif key == "优先级":
                priority = value.lower().replace("p", "").replace(" ", "")
                if priority.startswith("0"):
                    result["priority"] = "p0"
                elif priority.startswith("1"):
                    result["priority"] = "p1"
                elif priority.startswith("2"):
                    result["priority"] = "p2"
                elif priority.startswith("3"):
                    result["priority"] = "p3"
            elif key in ("测试类型",):
                result["test_type"] = value.lower()
            elif key in ("前置条件",):
                result["precondition"] = value
            elif key in ("状态",):
                result["status"] = value.lower()
            continue

        # 测试步骤段落
        if any(kw in lower for kw in ["测试步骤", "步骤", "操作步骤"]):
            in_steps = True
            in_expected = False
            continue
        if any(kw in lower for kw in ["预期结果", "期望结果", "预期"]):
            in_expected = True
            in_steps = False
            continue
        if in_steps and stripped:
            steps_lines.append(stripped)
        elif in_expected and stripped:
            expected_lines.append(stripped)

    if steps_lines:
        result["test_steps"] = "\n".join(steps_lines)
    if expected_lines:
        result["expected_result"] = "\n".join(expected_lines)

    # 如果行数少，把整块当 steps
    if not result.get("test_steps") and not result.get("expected_result"):
        if len(lines) >= 2:
            result["test_steps"] = "\n".join(lines[1:])

    return result


# ======================================================================
# Markdown 导出
# ======================================================================


def export_cases_to_markdown(cases: list) -> str:
    """导出测试用例列表为 Markdown 表格格式。

    返回包含表格和元数据的完整 Markdown 字符串。
    """
    lines = ["# 测试用例", "", f"**导出时间**: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}", f"**用例总数**: {len(cases)}", ""]

    for i, case in enumerate(cases, 1):
        name = case.name if hasattr(case, 'name') else (case.get('name', '') if isinstance(case, dict) else '')
        module = case.module if hasattr(case, 'module') else (case.get('module', '') if isinstance(case, dict) else '') or ''
        priority = (case.priority if hasattr(case, 'priority') else (case.get('priority', 'p2') if isinstance(case, dict) else 'p2')) or 'p2'
        test_type = (case.test_type if hasattr(case, 'test_type') else (case.get('test_type', 'functional') if isinstance(case, dict) else 'functional')) or 'functional'
        precondition = (case.precondition if hasattr(case, 'precondition') else (case.get('precondition', '') if isinstance(case, dict) else '')) or ''
        test_steps = (case.test_steps if hasattr(case, 'test_steps') else (case.get('test_steps', '') if isinstance(case, dict) else '')) or ''
        expected_result = (case.expected_result if hasattr(case, 'expected_result') else (case.get('expected_result', '') if isinstance(case, dict) else '')) or ''
        status = (case.status if hasattr(case, 'status') else (case.get('status', 'draft') if isinstance(case, dict) else 'draft')) or 'draft'

        lines.append(f"## TC{i:03d}：{name}")
        lines.append("")
        lines.append(f"| 字段 | 值 |")
        lines.append(f"|------|----|")
        lines.append(f"| 模块 | {module} |")
        lines.append(f"| 优先级 | {priority} |")
        lines.append(f"| 测试类型 | {test_type} |")
        lines.append(f"| 状态 | {status} |")
        lines.append("")
        if precondition:
            lines.append("### 前置条件")
            lines.append("")
            lines.append(precondition)
            lines.append("")
        if test_steps:
            lines.append("### 测试步骤")
            lines.append("")
            lines.append(test_steps)
            lines.append("")
        if expected_result:
            lines.append("### 预期结果")
            lines.append("")
            lines.append(expected_result)
            lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)


# ======================================================================
# XMind 导出（XMind 文件是 ZIP 包，内含 content.xml）
# ======================================================================


def export_cases_to_xmind(cases: list) -> bytes:
    """导出测试用例为 XMind 格式。

    构建符合 XMind 8/2020 规范的 ZIP 包：
    - META-INF/manifest.xml
    - meta.xml
    - content.xml（主题-子主题树结构）
    """
    import zipfile
    import io

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # META-INF/manifest.xml
        zf.writestr("META-INF/manifest.xml", """<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<manifest xmlns="urn:xmind:xmap:xmlns:manifest:1.0">
  <file-entry full-path="content.xml" media-type="text/xml"/>
  <file-entry full-path="meta.xml" media-type="text/xml"/>
</manifest>""")

        # meta.xml
        zf.writestr("meta.xml", f"""<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<meta xmlns="urn:xmind:xmap:xmlns:meta:2.0" version="2.0">
  <Creator>
    <Name>AI-HUB</Name>
    <Version>1.0</Version>
  </Creator>
  <Created>{__import__('datetime').datetime.now().isoformat()}</Created>
</meta>""")

        # content.xml — 构建用例思维导图
        topics_xml = ""
        for i, case in enumerate(cases, 1):
            name = _case_attr(case, 'name', '')
            if not name:
                continue
            priority = _case_attr(case, 'priority', 'p2')
            status = _case_attr(case, 'status', 'draft')
            module = _case_attr(case, 'module', '')
            label = f"{name} [{priority}]"
            topics_xml += f"""<topic id="{i}" structure-class="org.xmind.ui.logic.right">
          <title>{_xml_escape(label)}</title>
          <children>
            <topics type="attached">
              <topic id="{i}-note">
                <title>{_xml_escape(f'模块: {module} | 状态: {status}')}</title>
              </topic>
            </topics>
          </children>
        </topic>"""

        content = f"""<?xml version="1.0" encoding="UTF-8" standalone="no"?>
<xmap-content xmlns="urn:xmind:xmap:xmlns:content:2.0" xmlns:fo="http://www.w3.org/1999/XSL/Format" xmlns:svg="http://www.w3.org/2000/svg" version="2.0">
  <sheet id="sheet1">
    <title>测试用例</title>
    <topic id="root" structure-class="org.xmind.ui.logic.right">
      <title>测试用例（共{len(cases)}条）</title>
      <children>
        <topics type="attached">
          {topics_xml}
        </topics>
      </children>
    </topic>
  </sheet>
</xmap-content>"""
        zf.writestr("content.xml", content)

    buf.seek(0)
    return buf.getvalue()


def _case_attr(case, attr: str, default=""):
    """安全地从 ORM 对象或 dict 中读取属性"""
    if hasattr(case, attr):
        val = getattr(case, attr)
        return val if val is not None else default
    if isinstance(case, dict):
        val = case.get(attr)
        return val if val is not None else default
    return default


def _xml_escape(text: str) -> str:
    """XML 转义"""
    text = text.replace("&", "&amp;")
    text = text.replace("<", "&lt;")
    text = text.replace(">", "&gt;")
    text = text.replace('"', "&quot;")
    text = text.replace("'", "&apos;")
    return text


# ======================================================================
# CSV 导入解析
# ======================================================================


def parse_csv_cases(content: bytes) -> list[dict]:
    """解析 CSV 内容为用例字典列表。

    期望表头：用例标题, 模块, 优先级, 测试类型, 前置条件, 测试步骤, 预期结果, 状态, 标签
    """
    import csv
    import io

    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    cases = []

    for row in reader:
        title = _csv_val(row, '用例标题', '')
        if not title:
            continue

        priority = _csv_val(row, '优先级', 'p2').lower()
        if priority not in ('p0', 'p1', 'p2', 'p3'):
            priority = 'p2'

        status = _csv_val(row, '状态', 'draft').lower()
        if status not in ('draft', 'active', 'deprecated'):
            status = 'draft'

        test_type = _csv_val(row, '测试类型', 'functional').lower()

        tags_str = _csv_val(row, '标签', '')
        tags = [t.strip() for t in tags_str.split(',') if t.strip()] if tags_str else None

        cases.append({
            'name': title,
            'module': _csv_val(row, '模块') or None,
            'priority': priority,
            'test_type': test_type,
            'precondition': _csv_val(row, '前置条件') or None,
            'test_steps': _csv_val(row, '测试步骤') or None,
            'expected_result': _csv_val(row, '预期结果') or None,
            'status': status,
            'tags': tags,
        })

    return cases


def _csv_val(row: dict, key: str, default: str = '') -> str:
    """安全地从 CSV 行中取值"""
    val = row.get(key, row.get(key.strip(), default))
    return str(val).strip() if val else default


# ======================================================================
# XMind / FreeMind 导入解析
# ======================================================================


def parse_mindmap_cases(content: bytes) -> list[dict]:
    """解析 XMind (.xmind) 或 FreeMind (.mm) 文件中的用例。

    从 ZIP（XMind）或 XML（FreeMind）中提取主题文本，尝试按 Markdown 格式解析。
    """
    import xml.etree.ElementTree as ET

    # 尝试作为 XMind (ZIP) 读取
    try:
        import zipfile
        import io
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            if 'content.xml' in zf.namelist():
                xml_content = zf.read('content.xml')
                return _parse_xmind_xml(xml_content)
    except (zipfile.BadZipFile, Exception):
        pass

    # 尝试作为 FreeMind XML 直接读取
    try:
        return _parse_freemind_xml(content)
    except Exception:
        return []


def _parse_xmind_xml(xml_content: bytes) -> list[dict]:
    """从 XMind content.xml 中提取主题文本为用例"""
    import xml.etree.ElementTree as ET
    cases = []

    try:
        root = ET.fromstring(xml_content)
        # XMind XML namespace
        ns = {'x': 'urn:xmind:xmap:xmlns:content:2.0'}
        # 递归遍历所有 topic 元素
        for topic in root.iter():
            if topic.tag == '{urn:xmind:xmap:xmlns:content:2.0}topic':
                title_el = topic.find('x:title', ns)
                if title_el is not None and title_el.text:
                    title = title_el.text.strip()
                    if title:
                        case = {'name': title, 'priority': 'p2', 'status': 'active',
                                'test_type': 'functional', 'tags': None,
                                'module': None, 'precondition': None,
                                'test_steps': None, 'expected_result': None}
                        # 尝试从标题中提取优先级 [P0] [P1] 等
                        import re
                        prio_match = re.search(r'\[(P[0-3])\]', title)
                        if prio_match:
                            case['priority'] = prio_match.group(1).lower()
                            case['name'] = title.replace(f'[{prio_match.group(1)}]', '').strip()
                        cases.append(case)
    except Exception:
        pass

    return cases


def _parse_freemind_xml(content: bytes) -> list[dict]:
    """从 FreeMind (.mm) XML 中提取主题"""
    import xml.etree.ElementTree as ET
    cases = []

    try:
        root = ET.fromstring(content)
        for node in root.iter('node'):
            text = node.get('TEXT', '')
            if text and len(text) > 3:
                cases.append({
                    'name': text, 'priority': 'p2', 'status': 'active',
                    'test_type': 'functional', 'tags': None,
                    'module': None, 'precondition': None,
                    'test_steps': None, 'expected_result': None,
                })
    except Exception:
        pass

    return cases
